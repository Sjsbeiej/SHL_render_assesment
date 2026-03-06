"""Analyze at what rank each relevant assessment appears in the retriever's candidate pool."""
import openpyxl
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from graph import query_analyzer_node, retriever_node, GraphState

def normalize_url(url):
    url = url.strip().rstrip("/").lower()
    url = url.replace("/solutions/products/product-catalog/view/", "/products/product-catalog/view/")
    return url

def slug_from_url(url):
    return url.split("/view/")[-1].rstrip("/")

def process_query(qi, query, relevant_urls):
    """Run query analyzer + retriever for a single query."""
    state = GraphState(
        query=query, search_queries=[], skills=[], max_duration=None,
        domain="", candidates=[], recommendations=[]
    )
    qa_result = query_analyzer_node(state)
    state.update(qa_result)
    ret_result = retriever_node(state)
    return qi, query, relevant_urls, state, ret_result['candidates']

# Load train set
wb = openpyxl.load_workbook("input/Gen_AI Dataset.xlsx")
ws = wb["Train-Set"]
query_urls = defaultdict(list)
for row in range(2, ws.max_row + 1):
    q = ws.cell(row=row, column=1).value
    u = ws.cell(row=row, column=2).value
    if q and u:
        query_urls[q].append(u)

queries = list(query_urls.items())

# Run all queries in parallel (LLM + retrieval)
print("Running all 10 queries in parallel...")
results = {}
with ThreadPoolExecutor(max_workers=5) as executor:
    futures = {
        executor.submit(process_query, qi, query, urls): qi
        for qi, (query, urls) in enumerate(queries, 1)
    }
    for future in as_completed(futures):
        qi = futures[future]
        results[qi] = future.result()
        print(f"  Q{qi} done")

# Print results in order
total_found = 0
total_relevant = 0
for qi in sorted(results.keys()):
    qi, query, relevant_urls, state, candidates = results[qi]

    print(f"\n{'='*70}")
    print(f"Q{qi}: {query[:80]}...")
    print(f"Relevant: {len(relevant_urls)} assessments")
    print("="*70)

    print(f"\nSearch queries ({len(state['search_queries'])}):")
    for sq in state['search_queries']:
        print(f"  - {sq[:80]}")

    print(f"\nTotal candidates retrieved: {len(candidates)}")

    candidate_urls_norm = [normalize_url(c['url']) for c in candidates]
    relevant_norm = [normalize_url(u) for u in relevant_urls]

    all_found_by = None
    found_count = 0

    for rel_url, rel_orig in zip(relevant_norm, relevant_urls):
        slug = slug_from_url(rel_orig)
        found_at = None
        for rank, cand_url in enumerate(candidate_urls_norm, 1):
            if rel_url == cand_url:
                found_at = rank
                break
        if found_at:
            found_count += 1
            status = f"FOUND at rank {found_at}"
        else:
            status = "NOT IN POOL"
        print(f"  {slug:55s} => {status}")

    for k in [10, 20, 30, 40, 50, 60, 80, 100, 120]:
        top_k_urls = set(candidate_urls_norm[:k])
        hits = sum(1 for u in relevant_norm if u in top_k_urls)
        if hits == len(relevant_norm):
            all_found_by = k
            break

    print(f"\n  Found {found_count}/{len(relevant_urls)} in pool")
    if all_found_by:
        print(f"  ALL found within top {all_found_by} candidates")
    else:
        print(f"  NOT all found even in top {len(candidates)}")

    total_found += found_count
    total_relevant += len(relevant_urls)

print(f"\n{'='*70}")
print(f"TOTAL: {total_found}/{total_relevant} found in pool")
print("="*70)
