"""
Evaluation pipeline for the SHL Assessment Recommendation Engine.
Computes Recall@K on the train set and generates predictions for the test set.
"""

import csv
import os
from collections import defaultdict
from urllib.parse import urlparse

import openpyxl

import config
from graph import recommend


def normalize_url(url: str) -> str:
    """Normalize URL for comparison (remove trailing slash, standardize domain)."""
    url = url.strip().rstrip("/").lower()
    # Normalize both URL patterns
    url = url.replace("/solutions/products/product-catalog/view/",
                      "/products/product-catalog/view/")
    url = url.replace("www.shl.com/products/", "www.shl.com/products/")
    return url


def load_train_set(filepath: str) -> dict[str, list[str]]:
    """Load train set: {query: [list of relevant URLs]}."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Train-Set"]

    query_urls = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        query = ws.cell(row=row, column=1).value
        url = ws.cell(row=row, column=2).value
        if query and url:
            query_urls[query].append(normalize_url(url))

    return dict(query_urls)


def load_test_set(filepath: str) -> list[str]:
    """Load test set queries."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Test-Set"]

    queries = []
    for row in range(2, ws.max_row + 1):
        query = ws.cell(row=row, column=1).value
        if query:
            queries.append(query)

    return queries


def compute_recall_at_k(recommended_urls: list[str], relevant_urls: list[str], k: int = 10) -> float:
    """Compute Recall@K for a single query."""
    recommended_normalized = [normalize_url(u) for u in recommended_urls[:k]]
    relevant_normalized = [normalize_url(u) for u in relevant_urls]

    if not relevant_normalized:
        return 0.0

    hits = sum(1 for url in relevant_normalized if url in recommended_normalized)
    return hits / len(relevant_normalized)


def evaluate_train_set(dataset_path: str):
    """Evaluate on the train set and report Recall@K."""
    print("Loading train set...")
    query_urls = load_train_set(dataset_path)
    print(f"Loaded {len(query_urls)} queries with {sum(len(v) for v in query_urls.values())} total labels")

    recalls = []
    for i, (query, relevant_urls) in enumerate(query_urls.items(), 1):
        print(f"\n--- Query {i}/{len(query_urls)} ---")
        print(f"Query: {query[:80]}...")
        print(f"Relevant URLs: {len(relevant_urls)}")

        recommendations = recommend(query)
        recommended_urls = [r["url"] for r in recommendations]

        recall = compute_recall_at_k(recommended_urls, relevant_urls, k=config.TOP_K_FINAL)
        recalls.append(recall)

        print(f"Recommended: {len(recommended_urls)} assessments")
        print(f"Recall@{config.TOP_K_FINAL}: {recall:.4f}")

        # Show matches
        rec_normalized = [normalize_url(u) for u in recommended_urls]
        for url in relevant_urls:
            match = "HIT" if normalize_url(url) in rec_normalized else "MISS"
            name = url.split("/view/")[-1].rstrip("/")
            print(f"  [{match}] {name}")

    mean_recall = sum(recalls) / len(recalls) if recalls else 0.0
    print(f"\n{'=' * 60}")
    print(f"Mean Recall@{config.TOP_K_FINAL}: {mean_recall:.4f}")
    print(f"Per-query recalls: {[f'{r:.2f}' for r in recalls]}")
    return mean_recall


def generate_test_predictions(dataset_path: str, output_path: str):
    """Generate predictions CSV for the test set."""
    print("\nLoading test set...")
    queries = load_test_set(dataset_path)
    print(f"Loaded {len(queries)} test queries")

    rows = []
    for i, query in enumerate(queries, 1):
        print(f"\n--- Test Query {i}/{len(queries)} ---")
        print(f"Query: {query[:80]}...")

        recommendations = recommend(query)
        print(f"Generated {len(recommendations)} recommendations")

        for rec in recommendations:
            rows.append([query, rec["url"]])

    # Write CSV
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Query", "Assessment_url"])
        writer.writerows(rows)

    print(f"\nPredictions saved to {output_path}")
    print(f"Total rows: {len(rows)}")


if __name__ == "__main__":
    dataset_path = os.path.join(os.path.dirname(__file__), "input", "Gen_AI Dataset.xlsx")

    print("=" * 60)
    print("EVALUATION ON TRAIN SET")
    print("=" * 60)
    mean_recall = evaluate_train_set(dataset_path)

    print(f"\n\n{'=' * 60}")
    print("GENERATING TEST SET PREDICTIONS")
    print("=" * 60)
    output_path = os.path.join(os.path.dirname(__file__), "predictions.csv")
    generate_test_predictions(dataset_path, output_path)
